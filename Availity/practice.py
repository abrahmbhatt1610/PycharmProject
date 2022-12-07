# import os
#
#
# def file_name():
#     file1 = "C:/Users/HP/Downloads/CheckMultipleDetails-2022-10-22-05-20-27.pdf"
#     file2 = "C:/Users/HP/Downloads/AR0001897456 MEHTA HMO PARTNERS DBA HEALTH ADVANTAGE EOB.pdf"
#     os.rename(file1, file2)
#     print(file2)
#
import os

from Availity.Locators.LoginPageLoactors import *

os.listdir()
print(f'{os.listdir("C:/Users/HP/OneDrive - EMP Claim Solutions Inc/Documents/Test")}')
file1 = "C:\\Users\HP\\OneDrive - EMP Claim Solutions Inc\\Documents\\Test\\"
file2 = 'C:/Users/HP/Downloads/' + filename + '.pdf'
os.rename(file1, file2)
print("Successfully")
print(file1)
import openpyxl
from openpyxl import workbook

from pandas.core.interchange import column

# to load the workbook with its path
# bk = openpyxl.load_workbook('C:/Users/HP/Downloads/RemittanceViewer_CheckEftList_20221017041237.xlsx')
# to identify active worksht
# sht = bk.active
# c1 = sht.cell(row=1, column=2)

# to identify maximum rows count
# print(sht.max_row)
# to identify maximum columns count
# print(c1.value)
# for i in range(1, sht.max_row + 1):
#     # to iterate through the columns
#
#     for j in range(1, sht.max_column + 1):
#         # to get the cell data and
#         print(sht.cell(row=i, column=j).value)
# import pandas as pd
# import xlrd
#
# file = 'C:/Users/HP/Downloads/RemittanceViewer_CheckEftList_20221017041237 (1).xlsx'
# xl = pd.ExcelFile(file)
# print(xl.sheet_names)
#
# workbook = xlrd.open_workbook('C:/Users/HP/Downloads/RemittanceViewer_CheckEftList_20221017041237 (1)')
# workbook = xlrd.open_workbook('C:/Users/HP/Downloads/RemittanceViewer_CheckEftList_20221017041237 (1)',
#                               on_demand=True)
#
# worksheet = workbook.sheet_by_name('Sheet2')
# import PyPDF2
#
# file = open("C:/Users/HP/Downloads/RemittanceViewer_CheckEftList_20221017041237 (1).pdf", "rb")
#
# reader = PyPDF2.PdfFileReader(file)
#
# page1 = reader.getPage(0)
#
# pdfData = page1.extractText("Check/EFT Trace Number")
#
# print(pdfData)
#
# assert "Check/EFT Trace Number" in pdfData
#
# # assert "Mukesh" in pdfData
